# -*- encoding: utf-8 -*-
from openpyxl import load_workbook, Workbook

def compact_relation(input_file, output_file):
    workbook = load_workbook(input_file)
    booksheet = workbook.active # 当前活跃的sheet，默认第一个sheet
    relation_col = 7 # relation在xlsx文件的第几列
    # 行数
    row_num = booksheet.max_row

    ignore_relations = ['child', 'family_member', 'parent', 'relative', 'donor', 'great_great_grandson', 'master', 'pupil', 'neighbour']

    # 下标从1开始计数
    for i in range(2, row_num+1):
        # 跳过第一行
        relation = booksheet.cell(row=i, column=relation_col).value
        relation_changed = ''
        if relation in ['servant_girl']:
            relation_changed = 'servant'
        elif relation in ['nurser', 'adopted_mother']:
            relation_changed = 'mother'
        elif relation in ['daughter_in_law']:
            relation_changed = 'daughter'
        elif relation in ['concubine']:
            relation_changed = 'wife'
        elif relation in ['son_in_law', 'adopted_son']:
            relation_changed = 'son'
        elif relation in ['girl_friend']:
            relation_changed = 'friend'
        elif relation in ['elder_brother', 'elder_male_cousin', 'younger_brother', 'brother_in_law']:
            relation_changed = 'brother'
        elif relation in ['meme_ama']:
            relation_changed = 'father'
        elif relation in ['younger_sister']:
            relation_changed = 'sister'
        else:
            relation_changed = relation

        booksheet.cell(row=i, column=relation_col).value = relation_changed

    # 精简完关系，删除我们想忽略的上层关系
    for i in reversed(range(2, row_num+1)):
        relation = booksheet.cell(row=i, column=relation_col).value
        if relation in ignore_relations:
            booksheet.delete_rows(i)

    workbook.save(output_file)
